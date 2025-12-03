from django.shortcuts import render
from django.http import JsonResponse,HttpResponse
import io
from openpyxl import Workbook, load_workbook # Excel read/write
from django.contrib.auth import authenticate,login,logout


from core.auth import CsrfExemptSessionAuthentication
from .models import Project,Task,BOM
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status,viewsets
from .serializers import ProjectSerializer,TaskSerializer,BOMSerializer
from django.core.paginator import Paginator, EmptyPage
from django.db.models import Q,Count
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.authentication import BasicAuthentication
from rest_framework.parsers import MultiPartParser, FormParser

from django.http import HttpResponse




class ProjectDetail(APIView):
    def get(self, request, project_id):
        try:
            project = Project.objects.get(pk=project_id)
        except Project.DoesNotExist:
            return Response(
                {"error":"Project not found"},
                status= status.HTTP_404_NOT_FOUND
            )
        serializer = ProjectSerializer(project)
        return Response(serializer.data)




def home(request):
    # context = {"user_name":"Santh"}
    return render(request,"core/dashboard.html")

# Create your views here.

def project_tasks(request,project_id: int):
    try:
        project = Project.objects.get(pk=project_id)
    except Project.DoesNotExist:
        return JsonResponse({"error":"project not found"},status=404)
    data = [{
        "id":t.id,
        "title":t.title,
        "status": t.get_status_display(),
        "due_date":t.due_date.isoformat() if t.due_date else None,
        "created_at":t.created_at.isoformat(),
    } for t in project.tasks.all()
    ]
    return JsonResponse(
        {"project": project.name,"tasks detail":data}
    )


class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all().order_by("-created_at")
    serializer_class = ProjectSerializer

class TaskViewSet(viewsets.ModelViewSet):
    queryset = Task.objects.all().order_by("-created_at")
    serializer_class = TaskSerializer


class ProjectList(APIView):
    # only logged-in user can hit this end point
    authentication_classes = [CsrfExemptSessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated] 
    def get(self,request):
        projects = Project.objects.filter(owner=request.user).order_by("name")
        
        try:
            page = int(request.GET.get("page",1))
        except ValueError:
            page = 1
        try:
            page_size = int(request.GET.get("page_size",10))
        except ValueError:
            page_size = 10
        if page_size <= 0:
            page_size = 10
        paginator = Paginator(projects, page_size)
        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            page_obj = []
        
        serializer = ProjectSerializer(page_obj, many=True)

        data ={
            "result" : serializer.data,
            "page":page,
            "page_size":page_size,
            "total_pages":paginator.num_pages,
            "total_items":paginator.count,
        }

        return Response(data)
    



    
    def post(self,request):
        serializer = ProjectSerializer(data=request.data)

        if serializer.is_valid():
            # after validating it we are writing to the Db
            serializer.save(owner=request.user)
            return Response(
                serializer.data,
                status=status.HTTP_201_CREATED
            )
        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST
        )
        

class TaskList(APIView):
    # def get(self,request):
    #     # ORM: get all tasks from DB (still a lazy query)
    #     tasks = Task.objects.all()
    #     # many = True because this is a list of objects , not one object
    #     serializer = TaskSerializer(tasks,many=True)
    #     # serializer.data -> converts all task instances to python dicts
    #     return Response(serializer.data)

    authentication_classes = [CsrfExemptSessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self,request):
        qs = Task.objects.filter(project__owner=request.user)
        # Read query parameters from URL: /api/ver2/tasks/?project=1&status=
        project_id = request.query_params.get("project")
        status_code = request.query_params.get("status")
        priority = request.query_params.get("priority")
        sort = request.query_params.get("sort")
        
        if project_id is not None:
            qs = qs.filter(project_id=project_id)
        if status_code is not None:
            qs = qs.filter(status=status_code)
        
        if priority is not None:
            qs = qs.filter(priority=priority)
        if sort == "due":
            qs = qs.order_by("due_date")
        elif sort == "new":
            qs = qs.order_by("-created_at")
        else:
            qs = qs.order_by("title")
        #----- Search ('q') ------------------
        # use double under score after field name : title__icontains
        q = request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(title__icontains=q) |
                Q(status__icontains=q)
            )

        # ---  Pagination ------
        try: # - request.GET is a dictionary-like object 
            #containing all query parameters from the URL
            page = int(request.GET.get('page',1))
        except ValueError:
            page =1
        try:
            page_size = int(request.GET.get("page_size",10))
        except:
            page_size = 10
        paginator = Paginator(qs,page_size)

        try:
            page_obj = paginator.page(page)
        except EmptyPage:
            page_obj = []
        serilaizer = TaskSerializer(page_obj,many=True)
        data = {
            "results":serilaizer.data,
            "page":page,
            "page_size":page_size,
            "total_page":paginator.num_pages,
            "total_items":paginator.count,

        }
        return Response(data)

    def post(self,request):
        # DRF parses JSON body and puts it into request.data(a dict)
        serializer = TaskSerializer(data=request.data)
        # run the validation : type checks, req fields, choces etc..,
        if serializer.is_valid():
            project = serializer.validated_data.get("project")
            if project.owner != request.user:
                return Response({
                    "detail":"Cannot add projects tou don't own"
                }, status=status.HTTP_403_FORBIDDEN)

            task = serializer.save()

            return Response(TaskSerializer(task).data,
                            status=status.HTTP_201_CREATED)
        return Response(serializer.errors,
                        status=status.HTTP_400_BAD_REQUEST)
    
class TaskDetail(APIView):
    authentication_classes = [CsrfExemptSessionAuthentication,BasicAuthentication]
    permission_classes = [IsAuthenticated]

    # handle get,put,delete
    def get_object(self,task_id,user):
        try:
            task = Task.objects.get(pk=task_id,project__owner=user)
            return task
        except Task.DoesNotExist:
            return None
    def get(self,request,task_id):
        task = self.get_object(task_id,request.user)
        if task is None:
            return Response(
                {"error":"Task not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = TaskSerializer(task)
        return Response(serializer.data)
    def put(self,request,task_id):
        task = self.get_object(task_id,request.user)
        if task is None:
            return Response(
                {"error":"Task not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = TaskSerializer(task,data=request.data)
        if serializer.is_valid():
            project = serializer.validated_data.get("project",task.project)
            if project.owner != request.user:
                return Response({"detail":"cannot remove projects you don't own"},
                                status=status.HTTP_403_FORBIDDEN)

            serializer.save()
            return Response(serializer.data)
        return Response(
            serializer.errors,status=status.HTTP_400_BAD_REQUEST
        )
    def patch(self,request,task_id):
        """
        Partial Update (Only end field you want to change)
        """
        task = self.get_object(task_id,request.user)

        if task is None:
            return Response({
                "error":"Task is not found"
            }, status=status.HTTP_404_NOT_FOUND)
        # Partial update
        serializer = TaskSerializer(task,data=request.data, partial=True)
        if serializer.is_valid():
            project = serializer.validated_data.get("project",task.project)
            if project.owner != request.user:
                return Response({"detail":"cannot remove projects you don't own"},
                                status=status.HTTP_403_FORBIDDEN)
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors,status=status.HTTP_400_BAD_REQUEST)

    def delete(self,request,task_id):
        task = self.get_object(task_id,request.user)
        if task is None:
            return Response(
                {"error":"Task not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        task.delete()
        return Response(
            status=status.HTTP_204_NO_CONTENT
        )
    

class ProjectOverview(APIView):
    def get(self,request, project_id):
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({
                "detail":"Project not found"
            }, status=status.HTTP_404_NOT_FOUND)
        qs = Task.objects.filter(project=project)
        total_tasks = qs.count()

        status_counts = qs.values("status").annotate(count=Count("id"))
        by_status = {item["status"]:item["count"] for item in status_counts}

        for status_value in Task.Status.values:
            by_status.setdefault(status_value,0)

        # count by priority
        priority_counts = qs.values("priority").annotate(count=Count("id"))
        by_priority ={item["priority"]:item["count"] for item in priority_counts}

        for priority_value in Task.Priority.values:
            by_priority.setdefault(priority_value,0)
        
            data = {
                "project_id":project.id,
                "project_name":project.name,
                "total_tasks":total_tasks,
                "by_status":by_status,
                "by_priority":by_priority,
            }
        return Response(data)
    
class ProjectBOMList(APIView):
    """
    List or create BOM items for a single project.
    URL pattern: /api/ver2/projects/<project_id>/bom/
    """
    authentication_classes = [CsrfExemptSessionAuthentication,BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def get_project(self,project_id,user):
        """
        Helper method:
        Safely fetch a project that belongs to the current user.
        Returns None if not found or not owned by user.
        """
        try:
            return Project.objects.get(pk=project_id, owner=user)
        except Project.DoesNotExist:
            return None
    def get(self,request,project_id):
        """
        Helper method:
        Safely fetch a project that belongs to the current user.
        Returns None if not found or not owned by user.
        """
        project = self.get_project(project_id,request.user)
        if project is None:
            return Response({
                "detail": "Project not found"
            }, status=status.HTTP_404_NOT_FOUND)
        # Use related_name="bom_items" from BOM.project
        qs = project.bom_items.all().order_by("category","model")
        serializer = BOMSerializer(qs,many=True)
        return Response(serializer.data)
    def post(self,request,project_id):
        """
        POST /api/ver2/projects/<project_id>/bom/
        Create one BOM row for this project.
        """
        project = self.get_project(project_id,request.user)
        if project is None:
            return Response({
                "detail": "Project not found"
            }, status=status.HTTP_404_NOT_FOUND)
        # We ignore any "project" field from the client and force it to this project
        serializer = BOMSerializer(data=request.data)
        if serializer.is_valid():
            bom_item =serializer.save(project=project)
            return Response(BOMSerializer(bom_item).data,
                            status=status.HTTP_201_CREATED)
        return Response(serializer.errors,
                        status=status.HTTP_400_BAD_REQUEST)
    
class BOMItemDetail(APIView):
     
     
     """
     Retrieve or delete a single BOM item by id.
     URL pattern: /api/ver2/bom/<item_id>/
     """
     authentication_classes =[CsrfExemptSessionAuthentication,BasicAuthentication]
     permission_classes =[IsAuthenticated]

     def get_object(self,item_id,user):
         """
        Helper: fetch a BOM item that belongs to a project owned by 'user'.
        Returns None if not found or not owned by this user.
        """
         try:
            return BOM.objects.get(pk=item_id,project__owner=user)
         except BOM.DoesNotExist:
             return None
     def get(self, request, item_id):
         """
        Optional: GET /api/ver2/bom/<item_id>/
        Returns the BOM item details.
        """
         bom_item = self.get_object(item_id,request.user)
         if bom_item is None:
             return Response({"detail":"BOM item not found"},
                            status=status.HTTP_404_NOT_FOUND)
         serializer = BOMSerializer(bom_item)
         return Response(serializer.data)
     def delete(self, request, item_id):
         """
        DELETE /api/ver2/bom/<item_id>/
        Deletes the BOM item if user owns the parent project.
        """
         bom_item = self.get_object(item_id,request.user)
         if bom_item is None:
             return Response({"detail":"BOM item not found"},
                            status=status.HTTP_404_NOT_FOUND)
         bom_item.delete()
         return Response(status=status.HTTP_204_NO_CONTENT)
     
class BOMExportView(APIView):
    """
    Export BOM from a single project as an excel file.
    URL: GET /api/ver2/projects/<project_id>/bom/export/
    """
    authentication_classes = [CsrfExemptSessionAuthentication, BasicAuthentication]
    permission_classes =[IsAuthenticated]

    def get_project(self, project_id, user):
        try:
            return Project.objects.get(pk=project_id, owner=user)
        except Project.DoesNotExist:
            return None
        
    def get(self, request, project_id):
        project = self.get_project(project_id, request.user)
        if project is None:
            return Response({"detail":"Project not found"},
                            status=status.HTTP_404_NOT_FOUND, )
        qs = project.bom_items.all().order_by("category","model")

        # 1. Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title ="BOM"

        # 2 Header row - must match your BOM columns
        headers = ["Category", "Model", "Description", "Qty", "Param1", "Param2", "Price"]
        ws.append(headers)
        # 3) Data rows
        for item in qs:
            ws.append([
                item.category,
                item.model,
                item.description,
                item.qty,
                item.param1,
                item.param2,
                item.price,  # Decimal or None – openpyxl handles it
            ])
        
        # 4) Save to in-memory bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # 5) Build HTTP Response
        filename  = f"project_{project.id}_bom.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["content-Disposition"] = f'attachment: filename="{filename}"'
        return response
    

class BOMImportView(APIView):
    """
    Import BOM rows from an uploaded Excel (.xlsx) file into a project.
    URL: POST /api/ver2/projects/<project_id>/bom/import
    Body: multipart/form-data with a 'file' field.
    """
    authentication_classes =[CsrfExemptSessionAuthentication,BasicAuthentication]
    permission_classes =[IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_project(self, project_id, user):
        try:
            return Project.objects.get(pk=project_id, owner = user)
        except Project.DoesNotExist:
            return None
    def post(self, request, project_id):
        project = self.get_project(project_id, request.user)
        if project is None:
            return Response(
                {"detail": "Project not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response(
                {"detail":"No file uploaded (expected field name 'file')"},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            wb = load_workbook(filename=file_obj, data_only=True)
        except Exception:
            return Response(
                {"detail": "Could not read Excel file. Make sure it is a .xlsx file."},
                 status=status.HTTP_400_BAD_REQUEST
            )
        sheet = wb.active
        # Expect first row to be header:
        # Category | Model | Description | Qty | Param1 | Param2 | Price
        rows_iter = list(sheet.iter_rows(values_only=True))
        if not rows_iter:
            return Response({"detail": "Excel file is empty"},
                            status=status.HTTP_400_BAD_REQUEST)
        # Simple header validation (optional)
        expected_header = ["Category", "Model", "Description", "Qty", "Param1", "Param2", "Price"]
        # You can make this case-insensitive if you want; for now we assume exact match.
        if list(rows_iter[0][:7]) != expected_header:
            return Response(
                {
                    "detail": "Unexpected header row. Expected: "
                    + " | ".join(expected_header)
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        bom_items = []
        for row in rows_iter[1:]:
            if not row or all(v in (None, "", 0) for v in row):
                continue

            category = str(row[0] or "")
            model = str(row[1] or "")
            description = str(row[2] or "")
            qty = int(row[3] or 1)
            param1 = str(row[4] or "")
            param2 = str(row[5] or "")
            price = row[6] if row[6] is not None else None

            bom_items.append(
                BOM(
                    project=project,
                    category=category,
                    model=model,
                    description=description,
                    qty=qty,
                    param1=param1,
                    param2=param2,
                    price=price,
                )
            )

        if not bom_items:
            return Response({"detail": "No valid BOM rows found"},
                            status=status.HTTP_400_BAD_REQUEST)

        # Bulk insert
        BOM.objects.bulk_create(bom_items)

        return Response({"imported": len(bom_items)},
                        status=status.HTTP_201_CREATED)

class LoginView(APIView):

    """
    Log in a user using username & password.
    URL: POST /api/auth/login/
    Body (JSON): { "username": "...", "password": "..." }
    """

    authentication_classes = [CsrfExemptSessionAuthentication, BasicAuthentication]
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get("username")
        password = request.data.get("password")
        if not username and password:
            return Response({"detail":"username and pass required"},status=status.HTTP_400_BAD_REQUEST)
        user = authenticate(request,username = username, password=password)
        if user is None:
            return Response({"detail": "Invalid username or password."},
                status=status.HTTP_400_BAD_REQUEST,)
        # now we will create session  and sets the session id cookie
        login(request,user)

        return Response({"user_id":user.id,
                         "username":user.username,
                         "is_authenticated": True},
                         status=status.HTTP_200_OK)
class LogoutView(APIView):
    """
    Log out current user
    """
    authentication_classes=[CsrfExemptSessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        logout(request)
        return Response({"detail": "Logged out"}, status=status.HTTP_200_OK)
    
class CurrenUserView(APIView):

    authentication_classes=[CsrfExemptSessionAuthentication, BasicAuthentication]
    permission_classes = [AllowAny]

    def get(self,request):
        user = request.user
        if not user.is_authenticated:
            return Response({"authenticated": False},status=status.HTTP_400_BAD_REQUEST)
        return Response({
                "id": user.id,
                "username": user.username,
                "is_authenticated": True,
        },status=status.HTTP_200_OK )
    










     
         
         
    
         
             
     
     

        
            


    
    
































    

